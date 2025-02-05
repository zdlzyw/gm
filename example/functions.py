# coding:utf8
"""
一些通用功能
"""
import datetime
import os
import struct
import time

import openpyxl
import pandas as pd
import select
import websocket
from colorama import Fore
from openpyxl import load_workbook, Workbook

from Agame.common.common import Defines, PACKETS
from Agame.common.common.MyException import NoneException
from Agame.common.common.config import Config, PROTO_HEAD_SIZE, PROTO_HEAD_FORMAT
from Lib.protobuf import descriptor as descriptor_mod

FD = descriptor_mod.FieldDescriptor
# 这句没用，就是为了让PACKETS能用上，忽略吧
PACKETS.C2G_SayHi()
DOSEPRINT = True  # 是否输出信息


def ran_user_name(prefix: str = Config.prefix):
	"""
	根据当前时间戳生成一个 前缀+7位数字 的字符串，为了兼容洞洞的老脚本，优化后就可以删除了
	:param prefix: 角色命名前缀 默认使用config里的配置
	:return: str
	"""
	game_account_f = time.time()
	date_array = datetime.datetime.fromtimestamp(game_account_f)
	hms = date_array.strftime("%Y%m%d")
	game_account_int = int(game_account_f * 1000000)
	return prefix + hms[-1:] + str(game_account_int)[-6:]


def sendpacket(packet):
	"""
	向服务器发送协议包
	:param packet:协议包
	:return: 发送成功返回True，发送失败返回None
	"""

	sendData = packet.getdatastream()  # 获取包信息
	prints(packet.person['uid'],  packet.person['username'],packet.__class__.__name__,  'send:', packet.toDict())

	try:
		packet.person['socket'].send(sendData, opcode=websocket.ABNF.OPCODE_BINARY)
	except ConnectionAbortedError as e:
		prints(packet.person['uid'], packet.person['username'], 'send failed', str(e))
		packet.person['socket_closed'] = True
		return None
	except Exception as e:
		prints(packet.person['uid'], packet.person['username'], 'send failed', str(e))
		return None
	return True


def getpacket(person):
	"""
	收包，把收到的包全都存起来，一个个解析
	:param person:把机器人传进来，使用该机器人的socket连接进行收包
	:return:
	"""
	if len(person['buffer']) >= PROTO_HEAD_SIZE:  # 包头长度，在config中配置
		__data = struct.unpack(PROTO_HEAD_FORMAT, person['buffer'][0:PROTO_HEAD_SIZE])
		size = __data[0]  # 包体长度
		messageid = __data[1]  # 协议id
		try:
			if len(person['buffer']) >= size:  # 如果接收到的协议大于包体长度才处理
				packet = eval("PACKETS.%s(person)" % Defines.PACKET_DEFINE[str(messageid)])  # 生成协议包体
				packet.filldatafromstream(person['buffer'][PROTO_HEAD_SIZE:size])  # 把除了包头部分的信息解析
				prints(person['uid'], person['username'], "receive", Defines.PACKET_DEFINE[str(messageid)])
				try:
					prints(Fore.LIGHTGREEN_EX, person['uid'], '||', person['username'], '||',
					       time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), '||', "packet dump:", packet.toDict())
				except Exception as e:
					prints('Error', e)
					pass
				person['buffer'] = person['buffer'][size:]  # 从收到的数据流中把该协议信息剔除
				return packet  # 返回协议包
		except KeyError:
			pass

	instream = select.select([person['socket']], [], [], 0)[0]  # socket的接收套接字信息
	if instream:  # 有信息传入
		try:
			buf_size = person['socket'].recv()
			person['buffer'] = buf_size
		except websocket.WebSocketConnectionClosedException:
			buf_size = person['socket'].recv()
			person['buffer'] = buf_size
			raise
	return None


def handlepacket(packet):
	"""
	协议包初始化
	:param packet:
	"""
	# 放在这个地方，先解码，然后把packet实例化
	prints(packet.person['uid'], packet.person['username'], 'Handle packet: ' + packet.__class__.__name__)
	packet.handle()


def handleinputstream(person):
	"""
	处理数据流
	:param person:收取数据的机器人
	"""
	while 1:
		packet = getpacket(person)
		if packet is None:
			break
		packet.handle()


def heartbeat(person):
	"""
	心跳包
	:param person:对应的机器人
	"""
	if time.time() - person['heartbeattime'] > 15:
		person['heartbeattime'] = time.time()
		person.MSG_C2G_KeepAlive()  # 心跳包


def waitforpacket(person, packetname, timeout=5):
	"""
	等待回包
	:param person:需要执行操作的机器人
	:param packetname:等待的包名
	:param timeout:设置超时时间，默认10秒
	:return:等到包体则返回packet，未等到包体则返回None
	"""
	prints(person['uid'], person['username'], 'Wait for packet: ' + str(packetname))
	latesttime = time.time()
	return_list = {}
	while 1:
		packet = getpacket(person)  # 收包
		if packet is None:  # 没收到
			total_time = int((time.time() - latesttime) * 1000)
			if total_time > timeout * 1000:  # 超时
				break
			else:
				time.sleep(0)
				continue
		if isinstance(packetname, str) and packet.__class__.__name__ == packetname:  # 收到包了，跟要等待的包做个比对
			handlepacket(packet)  # 是要等的包就处理一下，返回包体
			return packet
		elif isinstance(packetname, list) and packet.__class__.__name__ in packetname:
			return_list[packet.__class__.__name__] = packet
			if packet is None or ('ret' in packet.toDict() and packet['ret'] != 1):
				if len(return_list) < len(packetname):
					set1 = set(packetname)
					set2 = set(return_list.keys())
					for packet_name in list(set1 - set2):
						return_list[packet_name] = None
			if len(return_list) == len(packetname):
				return return_list
		elif packet.__class__.__name__ == 'S2C_CmdRet':
			return None
		else:  # 不是想等的包，处理一下，放到一个字典里，方便后续使用，但最多每一个协议名只存最新的20条协议信息
			handlepacket(packet)
			if 'C2G_Login' in person['packet_list'].keys() and len(person['packet_list']['C2G_Login']) > 0:
				_result = person['packet_list']['C2G_Login'].pop()
				person['packet_list']['C2G_Login'] = []
				if _result is not None and _result['ret'] in [2, 5, 6, 9]:
					packet.person['socket_closed'] = True
			if str(packet.__class__.__name__) in person['packet_list'].keys():
				if len(person['packet_list'][str(packet.__class__.__name__)]) >= 20:  # 20条之后先删除最老的包，然后再新增最新的包
					del person['packet_list'][str(packet.__class__.__name__)][0]
				person['packet_list'][str(packet.__class__.__name__)].append(packet)
			else:  # 空的就直接写入生成一个列表
				person['packet_list'][str(packet.__class__.__name__)] = [packet, ]
		heartbeat(person)
	prints(person['uid'], person['username'], 'Failed get packet: ' + str(packetname),
	       'use time {} ms'.format(total_time), None)
	return None


_accountdic = {}  # 用于记录当前机器人登录账号的角色的数字部分


def resetAccount(k=None):
	"""
	对_accountdic进行重置
	:param k:对某个key进行重置，默认为None，将# 对_accountdic置空
	:return:
	"""
	global _accountdic
	if k is None:
		_accountdic = {}
	else:
		if k in _accountdic.keys():
			_accountdic.pop(k)


def getUserAccount(server: str, _prefix: str = None, _start_with: int = None):
	"""
	用于生成序列账号
	:param server:服务器标识
	:param _prefix:前缀，默认为None，读取配置中的prefix
	:param _start_with:起始账号数字，默认为None，读取配置中的accountstartwith
	:return: 根据传入的服务器作为key，在该服下每创建一次，则更新一次数字，每生成一个账号数字+1
	"""
	if _start_with is None:
		accountstartwith = Config.account_startWith  # 起始账号数值（账号等于prefix+数字）
	else:
		try:
			accountstartwith = int(_start_with)
		except Exception:
			accountstartwith = 1
	if _prefix is None:
		prefix = Config.prefix  # 起始账号数值（账号等于prefix+数字）
	else:
		prefix = str(_prefix)
	global _accountdic
	if server in _accountdic.keys():
		_accountdic[server] += 1
	# if _accountdic[server] in [69, 93]:  # 屏蔽数字，跳过
	#     _accountdic[server] += 1
	else:
		_accountdic[server] = accountstartwith
	return prefix + str(_accountdic[server])


TRIGGER_INDEX = False


def is_trigger_start():
	"""
	用于控制机器人登录成功后执行操作的控制器
	:return: 可以开始后返回True，否则返回False
	"""
	global TRIGGER_INDEX
	if not TRIGGER_INDEX:
		tempfile = os.path.join(Config.loadTest_baseDir, 'trigger.txt')
		try:
			outputfile = open(tempfile.replace('\\', '/'), 'r')
			triggertext = outputfile.readline()
			outputfile.close()
			if triggertext == 'start':
				TRIGGER_INDEX = True
				return True
			else:
				return False
		except IOError:
			return False
	else:
		return True


def pb2dict(obj):
	"""
	protobuf转字典
	:param obj:proto协议对象
	:return: dict 协议信息
	"""
	adict = {}
	if not obj.IsInitialized():
		return None
	for field in obj.DESCRIPTOR.fields:
		if not getattr(obj, field.name):
			continue
		if not field.label == FD.LABEL_REPEATED:
			if not field.type == FD.TYPE_MESSAGE:
				adict[field.name] = getattr(obj, field.name)
			else:
				value = pb2dict(getattr(obj, field.name))
				if value:
					adict[field.name] = value
		else:
			if field.type == FD.TYPE_MESSAGE:
				adict[field.name] = [pb2dict(v) for v in getattr(obj, field.name)]
			else:
				adict[field.name] = {v for v in getattr(obj, field.name)}
	return adict


def parse_list(values, message):
	"""
	将列表转为protobuf
	:param values:数据
	:param message:协议
	"""
	if values:
		if isinstance(values[0], dict):  # value needs to be further parsed
			for v in values:
				cmd = message.add()
				parse_dict(v, cmd)
		else:  # value can be set
			message.extend(values)


def parse_dict(values, message):
	"""
	字典转protobuf
	:param values:数据
	:param message:协议
	"""
	for k, v in values.items():
		if isinstance(v, dict):  # 如果是字典或者列表，递归调用
			parse_dict(v, getattr(message, k))
		elif isinstance(v, list):
			parse_list(v, getattr(message, k))
		else:  # 如果是值则设值
			try:
				setattr(message, k, v)
			except AttributeError:
				prints('try to access invalid attributes %r.%r = %r', message, k, v)


def dict_to_protobuf(value, message):
	"""
	字典转protobuf
	:param value:数据
	:param message:协议
	"""
	parse_dict(value, message)


def response_time(func):
	"""
	装饰器，用于记录相应时间
	:param func:被装饰的方法名
	:return:
	"""

	def _reponse_time(*args, **kwargs):
		start_time = time.time() * 1000
		_result = func(*args, **kwargs)
		end_time = time.time() * 1000
		return _result, int(end_time - start_time)

	return _reponse_time


def get_exp_of_level(cur_level=1, level=200):
	"""
	获取升到level所需的经验数量
	:param cur_level: 当前等级
	:param level: 升到多少级
	:return:
	"""
	ex_file = os.path.join(Config.common_tab, 'role_info1.xlsx')
	wb = load_workbook(ex_file)
	ws = wb['Sheet1']
	maxrow = ws.max_row
	exp1 = 0
	exp2 = 0
	for _i in range(1, maxrow + 1):
		if ws.cell(_i, 1).value == cur_level:
			exp1 = ws.cell(_i, 3).value
		if ws.cell(_i, 1).value == level:
			exp2 = ws.cell(_i, 3).value
			wb.close()
			break
	return exp2 - exp1


def get_fuben_id(fuben_id: int):
	"""
	获取关卡为fuben_id的实际副本id
	:param fuben_id: 副本关卡
	:return:
	"""
	ex_file = os.path.join(Config.common_tab, 'role_info1.xlsx')
	wb = load_workbook(ex_file)
	ws = wb['fuben_info']
	maxrow = ws.max_row
	for _i in range(1, maxrow + 1):
		if ws.cell(_i, 1).value == fuben_id:
			only_id = ws.cell(_i, 2).value
			wb.close()
			return only_id


class ExcelHandle:
	def __init__(self, ex_file=None):
		"""新增了武将资质字段"""
		if not ex_file:
			ex_file = 'Agame道具类型表.xlsx'
		file_name = os.path.join(Config.common_tab, ex_file)  # 表格写死在common/info_table目录下的文件
		if not os.path.exists(file_name):
			raise NoneException('No such file: {}'.format(file_name))
		self.xl = openpyxl.load_workbook(file_name)

	def get_name_from_type_id(self, _type, _id):
		"""
		根据传入的道具type和id，查询道具的名字
		:param _type: type
		:param _id: id
		:return: str  resource_name or None
		"""

		table = self.xl.worksheets[0]
		col = table.max_row
		resource_name = f'{_type}-{_id}'
		for i in range(1, col + 1):
			if int(_type) == table.cell(i, 2).value and int(_id) == table.cell(i, 3).value:
				resource_name = table.cell(i, 1).value
		return resource_name

	def get_quality_from_type_id(self, _type, _id):
		"""
		根据传入的道具type和id，查询道具的名字
		:param _type: type
		:param _id: id
		:return: str  resource_name or None
		"""

		table = self.xl.worksheets[0]
		col = table.max_row
		resource_quality = f'{_type}-{_id}'
		for i in range(1, col + 1):
			if int(_type) == table.cell(i, 2).value and int(_id) == table.cell(i, 3).value:
				resource_quality = table.cell(i, 4).value
		return resource_quality

	def __del__(self):
		self.xl.close()


def write_to_excel(result_list, file_name='概率测试结果', server=None, write_file=False, item_info=None):
	"""
	对测试结果进行处理，把每个机器人抽到的结果单独列出一个表格，然后在第一个表格汇总概率结果
	:param result_list:  测试结果列表
	:param file_name:  写入文件名字
	:param server:  哪个服务器
	:param write_file:  是否写入文件
	:param item_info:  读取指定道具id映射表
	:return:
	"""
	prints('开始处理测试结果列表')
	wb = None
	ws = None
	huizong = None
	if write_file is True:
		num = 1
		while True:
			temp_file_name = f"{file_name}_{time.strftime('%Y%m%d', time.localtime())}_{num}.xlsx"  # 生成excel文件名
			temp_file_name = temp_file_name.replace('.txt', '')
			temp_file_name = os.path.join(os.path.expanduser('~'), "Desktop", temp_file_name)  # 生成最终文件名
			if os.path.exists(temp_file_name):  # 已有该文件，则数字+1
				num += 1
			elif num > 1000:
				return
			else:
				file_name = temp_file_name
				break
		wb = Workbook()  # 创建一个工作表
		huizong = wb.create_sheet('汇总结果')

	xl = ExcelHandle(item_info)  # 打开道具表，对照id-道具名
	rewards_dict = {}  # 用来存放抽到结果的道具信息
	final_result = {}  # 用来存放抽到道具的计数
	# 将每个账号的抽取结果写入单个文件表
	for user in result_list.keys():  # 遍历执行了几个账号
		if write_file is True:
			ws = wb.create_sheet(str(user))  # 生成一个user名命名的sheet
			lines = [['账号：', user, '服务器：', server], [], ['抽奖次数', '奖励']]  # 首行，显示一些配置信息及标题等
			for line in lines:  # 将标题等写入表格
				ws.append(line)
		final_result = {}  # 用来存放抽到道具的计数
		line_nums = list(result_list[user].keys())
		line_nums.sort()
		do_num = len(line_nums)
		for line_num in line_nums:
			line = []
			if write_file is True:
				line = ['第{}次'.format(line_num)]
			for ward in result_list[user][line_num]:
				if ward not in rewards_dict.keys():
					item = ward.split('-')
					if item[0] == '事件':
						rewards_dict[ward] = ward
					else:
						res_name = xl.get_name_from_type_id(item[0], item[1])
						if res_name is not None:
							rewards_dict[ward] = "{}-{}".format(res_name, item[2])
						else:
							rewards_dict[ward] = ward
				if write_file is True:
					line.append(rewards_dict[ward])
				if rewards_dict[ward] in final_result.keys():
					final_result[rewards_dict[ward]] += 1
				else:
					final_result[rewards_dict[ward]] = 1
			if write_file is True:
				ws.append(line)
		if write_file is True:
			# 汇总结果处理
			maxcol = huizong.max_column  # 当前最大列数
			if maxcol == 1:  # 如果是第一个就直接写入
				col_num = 1
			else:  # 不是第一个的话中间空一列出来
				col_num = maxcol + 2
			lines = [['账号：', user], ['服务器：', server], ['执行次数', do_num], [], ['汇总结果'],
			         ['道具名', '数量', '抽到次数', '概率']]  # 首行，显示一些配置信息及标题等
			row_num = 1  # 从第1行开始写入
			for line in lines:
				for word in line:  # 按照一行显示的内容写入数据
					huizong.cell(row_num, col_num + line.index(word), word)  # 写入数据
				row_num += 1  # 写完一行行号+1
			total_nums = sum(final_result.values())
			for k, v in final_result.items():
				temp_item = k.split('-')
				item_name = ''
				item_num = 0
				if len(temp_item) == 2:
					item_name = temp_item[0]
					item_num = temp_item[1]
				elif len(temp_item) == 3:
					item_name = '{}-{}'.format(temp_item[0], temp_item[1])
					item_num = temp_item[2]
				huizong.cell(row_num, col_num, item_name)  # 道具名写入
				huizong.cell(row_num, col_num + 1, item_num)  # 道具数量写入
				huizong.cell(row_num, col_num + 2, v)  # 抽取次数写入
				huizong.cell(row_num, col_num + 3, '{:.2%}'.format(v / total_nums))  # 概率计算写入
				row_num += 1

	if write_file is True:
		del wb['Sheet']  # 删除默认自带sheet
		wb.save(file_name)  # 保存文件
		prints('测试执行完毕，结果请到概率测试结果目录查找，{}'.format(file_name))
	return final_result


def prints(*args, **kwargs):
	""""""
	global DOSEPRINT
	if DOSEPRINT is True:
		args = [str(arg) for arg in args]
		print(' '.join(args))
		for k, v in kwargs.items():
			print(f'{k}={v}')


def get_server_time(timestr, offset):
	"""将时间戳+偏移量并返回"""
	timestr += offset
	return time.localtime(timestr)


def get_excel_info(excel_file, area=None, version=None, **kwargs):
	"""
	读取excel_file，并返回对应条件下的数据
	:param excel_file: excel文件名
	:param area: 地区
	:param version: 版本
	:param kwargs:
	:return:
	"""
	if area is None:
		area = 'Agame'
	if version is None:
		version = max([t_dir for t_dir in os.listdir(Config.configData_baseDir)
		               if os.path.isdir(os.path.join(Config.configData_baseDir, t_dir))])

	confDir = os.path.join(Config.configData_baseDir, version)
	if Config.Infomation[area]['area'] is None:
		areaDir = confDir
	else:
		areaDir = os.path.join(confDir, Config.Infomation[area]['area'])

	if not excel_file.endswith('.xlsx'):  # 不带后缀的自动补全
		excel_file = f'{excel_file}.xlsx'
	file_path = os.path.join(areaDir, excel_file)
	if not os.path.exists(file_path):  # 如果地区目录中存在该表，则读取之，不存在则读取国内版本的
		file_path = os.path.join(confDir, excel_file)
	if os.path.dirname(file_path) == Config.configData_baseDir:
		return []
	if not os.path.exists(file_path):
		print(f'不存在{file_path}文件，请检查文件名是否正确')
		return []

	df = pd.read_excel(file_path, sheet_name=0, header=4)
	condtions = ''
	if len(kwargs) > 0:
		for k, v in kwargs.items():
			if k not in list(df):
				continue
			if condtions is '':
				condtions = f'(df["{k}"] == {v})'
			else:
				condtions += f' & (df["{k}"] == {v})'
		df = eval(f'df[{condtions}]')
	return [dict(zip(list(df), d)) for d in df.values]


# 生成表格数据
def save_result(data_list, filename):
	"""
	data_list 数据列表
	filename 文件名称
	"""
	wb = Workbook()
	ws = wb.create_sheet('测试结果', 0)
	ws.append(['道具名称', '获得次数', '概率'])
	result_dic = {awardname: data_list.count(awardname) for awardname in data_list}
	for k,v in result_dic.items():
		ws.append([k,v,v/len(data_list)])
	wb.save(f'{filename}.xlsx')